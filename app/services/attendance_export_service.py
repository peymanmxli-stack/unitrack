"""
attendance_export_service.py

This service is responsible for exporting attendance data.

Teaching idea:
We do NOT want export logic directly inside routes.

Why?
Because routes should stay clean.

Clean architecture:
Route -> Service -> File generation

This file now supports:
- CSV export
- Excel export (.xlsx)

Professional idea:
Both export formats should use the SAME data preparation logic.
That way:
- CSV stays consistent
- Excel stays consistent
- future changes happen in one place only
"""

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def build_attendance_export_rows(attendance_records):
    """
    Build a reusable table-like data structure.

    Returns:
    A list of rows where:
    - first row = headers
    - remaining rows = attendance data

    Teaching idea:
    We centralize the export content here so both
    CSV and Excel use the exact same values.
    """

    rows = [
        [
            "Record ID",
            "User ID",
            "Username",
            "Full Name",
            "Role",
            "Movement Type",
            "Attendance Date",
            "Movement Time",
            "Notes",
            "Created At",
        ]
    ]

    for record in attendance_records:
        user = record.user

        full_name = ""
        if user:
            first_name = user.first_name or ""
            last_name = user.last_name or ""
            full_name = f"{first_name} {last_name}".strip()

        rows.append([
            record.id,
            record.user_id,
            user.username if user else "",
            full_name,
            user.role if user else "",
            record.movement_type,
            str(record.attendance_date) if record.attendance_date else "",
            record.movement_time.isoformat() if record.movement_time else "",
            record.notes if record.notes else "",
            record.created_at.isoformat() if record.created_at else "",
        ])

    return rows


def export_attendance_to_csv(attendance_records):
    """
    Convert attendance records into CSV text.

    Parameters:
    attendance_records -> list of Attendance model objects

    Returns:
    A string containing CSV data.

    Why StringIO?
    Because it lets Python build a file in memory
    without saving a physical file on disk first.
    """

    rows = build_attendance_export_rows(attendance_records)

    output = io.StringIO()
    writer = csv.writer(output)

    for row in rows:
        writer.writerow(row)

    csv_data = output.getvalue()
    output.close()

    return csv_data


def export_attendance_to_excel(attendance_records):
    """
    Convert attendance records into an Excel file in memory.

    Returns:
    Bytes data for a .xlsx file

    Why BytesIO?
    Because Excel files are binary files, not text files.

    Professional formatting added:
    - sheet title
    - bold header
    - dark header background
    - centered header text
    - auto-sized columns
    """

    rows = build_attendance_export_rows(attendance_records)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance Report"

    # Write all rows into the worksheet
    for row in rows:
        worksheet.append(row)

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="111827", end_color="111827")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Auto-fit column widths
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # Build binary Excel data in memory
    output = io.BytesIO()
    workbook.save(output)
    excel_data = output.getvalue()
    output.close()

    return excel_data